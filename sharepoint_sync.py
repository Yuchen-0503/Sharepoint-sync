"""
SharePoint Excel 数据同步脚本（单 Playwright session 版）
依赖: pip install playwright openpyxl
      playwright install chromium
"""

import io
import os
import time
import base64
import tempfile

import openpyxl
from playwright.sync_api import sync_playwright, Page

# ── 配置 ──────────────────────────────────────────────────────────────────────
SHAREPOINT_SITE = "https://wiscontrolchina.sharepoint.cn/sites/WISShanghai2"

FILE1_GUID = "6A41602F-E95D-4FB7-B175-01F7E9C0A736"
FILE2_GUID = "712640F7-BF87-40A4-8A78-C14CC6B5FB9D"

SHEET1_NAME      = "Results-2025"
TARGET_SHEET     = "进行中2026年"
SHEET_REPORTED   = "已出报告"
SHEET_JOINT      = "共同分级"
SHEET_JOINT_DONE = "共同分级已完成"

NO_MOVE_VALUE = "Supervision of Weighing & SDT"

# ── 列索引（1-based） ─────────────────────────────────────────────────────────
def col(letter: str) -> int:
    r = 0
    for ch in letter.upper():
        r = r * 26 + (ord(ch) - ord('A') + 1)
    return r

COL_A  = col("A")
COL_B  = col("B")
COL_H  = col("H")
COL_I  = col("I")
COL_J  = col("J")
COL_K  = col("K")
COL_P  = col("P")
COL_V  = col("V")
COL_AA = col("AA")

# ── Excel 辅助 ────────────────────────────────────────────────────────────────

def last_data_row(ws) -> int:
    last = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                last = cell.row
    return last

def get_row(ws, r: int) -> list:
    return [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

def set_row(ws, r: int, vals: list):
    for c, v in enumerate(vals, 1):
        ws.cell(row=r, column=c).value = v

def clear_row(ws, r: int):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).value = None

def append_rows(ws, rows: list):
    start = last_data_row(ws) + 1
    for i, row in enumerate(rows):
        set_row(ws, start + i, row)

# ── 登录 ─────────────────────────────────────────────────────────────────────

def login(page: Page):
    print("  打开 SharePoint 站点 …")
    page.goto(SHAREPOINT_SITE, timeout=60000)

    # 等待用户名输入框
    page.wait_for_selector(
        "input[type='email'], input[name='loginfmt'], input[name='username']",
        timeout=30000
    )
    print("  填写用户名 …")
    page.locator("input[type='email'], input[name='loginfmt'], input[name='username']").first.fill(USERNAME)

    # 找到并点击下一步（等按钮可用）
    next_btn = page.locator("input[type='submit'][id='idSIButton9'], button[type='submit']").first
    next_btn.wait_for(state="visible", timeout=10000)
    # 等按钮启用（页面 JS 验证邮箱格式后会启用）
    page.wait_for_timeout(1500)
    next_btn.click()

    # 等密码框
    page.wait_for_selector("input[type='password'], input[name='passwd']", timeout=30000)
    print("  填写密码 …")
    page.locator("input[type='password'], input[name='passwd']").first.fill(PASSWORD)
    page.wait_for_timeout(500)

    submit_btn = page.locator("input[type='submit'], button[type='submit']").first
    submit_btn.wait_for(state="visible", timeout=10000)
    page.wait_for_timeout(1000)
    submit_btn.click()

    # 处理「保持登录」提示
    try:
        page.wait_for_selector("input#idBtn_Back, button#declineButton", timeout=10000)
        print("  点击「不保持登录」…")
        page.locator("input#idBtn_Back, button#declineButton").first.click()
    except Exception:
        pass

    # 等待最终跳转到 SharePoint 站点（轮询 URL）
    for _ in range(30):
        time.sleep(2)
        if "sharepoint.cn" in page.url and "login" not in page.url:
            break
    print(f"  登录成功，URL: {page.url}")


# ── SharePoint API（在浏览器内执行） ─────────────────────────────────────────

def sp_api_get(page: Page, path: str) -> dict:
    url = f"{SHAREPOINT_SITE}/{path.lstrip('/')}"
    result = page.evaluate("""async (url) => {
        const r = await fetch(url, {headers: {'Accept': 'application/json;odata=verbose'}});
        if (!r.ok) throw new Error('HTTP ' + r.status + ' ' + url);
        return r.json();
    }""", url)
    return result


def get_file_server_url(page: Page, guid: str) -> str:
    data = sp_api_get(page, f"_api/web/GetFileById('{guid}')?$select=ServerRelativeUrl")
    return data["d"]["ServerRelativeUrl"]


def download_file_bytes(page: Page, server_relative_url: str) -> bytes:
    from urllib.parse import quote
    encoded = quote(server_relative_url)
    url = f"{SHAREPOINT_SITE}/_api/web/GetFileByServerRelativePath(decodedurl=@p)/$value?@p='{encoded}'"
    print(f"  下载: {server_relative_url.split('/')[-1]}")
    b64 = page.evaluate("""async (url) => {
        const r = await fetch(url);
        if (!r.ok) throw new Error('HTTP ' + r.status);
        const buf = await r.arrayBuffer();
        const bytes = new Uint8Array(buf);
        let binary = '';
        for (let i = 0; i < bytes.length; i++) binary += String.fromCharCode(bytes[i]);
        return btoa(binary);
    }""", url)
    return base64.b64decode(b64)


def upload_file_bytes(page: Page, server_relative_url: str, data: bytes):
    from urllib.parse import quote
    folder_url = "/".join(server_relative_url.split("/")[:-1])
    filename = server_relative_url.split("/")[-1]
    encoded_folder = quote(folder_url)
    upload_url = (
        f"{SHAREPOINT_SITE}/_api/web/GetFolderByServerRelativePath"
        f"(decodedurl=@p)/Files/Add(url='{filename}',overwrite=true)"
        f"?@p='{encoded_folder}'"
    )
    data_b64 = base64.b64encode(data).decode()
    print(f"  上传: {filename} ({len(data):,} bytes)")

    status = page.evaluate("""async ({digestUrl, uploadUrl, dataB64}) => {
        // 获取 digest
        const dr = await fetch(digestUrl, {
            method: 'POST',
            headers: {'Accept': 'application/json;odata=verbose'}
        });
        const dj = await dr.json();
        const digest = dj.d.GetContextWebInformation.FormDigestValue;

        // 上传
        const binary = atob(dataB64);
        const bytes = new Uint8Array(binary.length);
        for (let i = 0; i < binary.length; i++) bytes[i] = binary.charCodeAt(i);
        const r = await fetch(uploadUrl, {
            method: 'POST',
            headers: {
                'Accept': 'application/json;odata=verbose',
                'X-RequestDigest': digest,
                'Content-Type': 'application/octet-stream'
            },
            body: bytes
        });
        const text = await r.text();
        return {status: r.status, body: text.substring(0, 200)};
    }""", {
        "digestUrl": f"{SHAREPOINT_SITE}/_api/contextinfo",
        "uploadUrl": upload_url,
        "dataB64": data_b64,
    })

    print(f"  上传状态: {status['status']}")
    if status["status"] not in (200, 201):
        raise RuntimeError(f"上传失败: {status}")


# ── 数据处理 ─────────────────────────────────────────────────────────────────

def process(wb1: openpyxl.Workbook, wb2: openpyxl.Workbook) -> openpyxl.Workbook:
    ws1           = wb1[SHEET1_NAME]
    ws_target     = wb2[TARGET_SHEET]
    ws_reported   = wb2[SHEET_REPORTED]
    ws_joint      = wb2[SHEET_JOINT]
    ws_joint_done = wb2[SHEET_JOINT_DONE]

    # 步骤 1：筛选 K 列 == QINGDAO
    print("步骤 1: 筛选 K 列 == QINGDAO …")
    qingdao_rows = []
    for r in range(1, ws1.max_row + 1):
        k_val = ws1.cell(row=r, column=COL_K).value
        if k_val is not None and str(k_val).strip().upper() == "QINGDAO":
            qingdao_rows.append(get_row(ws1, r))
    print(f"  找到 {len(qingdao_rows)} 行")

    # 步骤 2：粘贴到 进行中2026年 J 列起始
    print("步骤 2: 追加到「进行中2026年」J 列 …")
    insert_start = last_data_row(ws_target) + 1
    for i, src in enumerate(qingdao_rows):
        dest_row = insert_start + i
        for j, val in enumerate(src):
            ws_target.cell(row=dest_row, column=COL_J + j).value = val

    total = ws_target.max_row

    # 步骤 3：AA→H，P→I（跳过表头）
    print("步骤 3: AA→H，P→I …")
    for r in range(2, total + 1):
        ws_target.cell(row=r, column=COL_H).value = ws_target.cell(row=r, column=COL_AA).value
        ws_target.cell(row=r, column=COL_I).value = ws_target.cell(row=r, column=COL_P).value

    # 步骤 4：按 H 列升序排序（表头固定）
    print("步骤 4: 按 H 列升序排序 …")
    data_rows = [get_row(ws_target, r) for r in range(2, ws_target.max_row + 1)]
    data_rows.sort(key=lambda row: (
        row[COL_H - 1] is None,
        str(row[COL_H - 1]) if row[COL_H - 1] is not None else ""
    ))
    for i, row_vals in enumerate(data_rows, start=2):
        clear_row(ws_target, i)
        set_row(ws_target, i, row_vals)

    # 步骤 5：V 列（非指定值）→ B 列
    print("步骤 5: V→B …")
    for r in range(2, ws_target.max_row + 1):
        v = ws_target.cell(row=r, column=COL_V).value
        if v is not None and str(v).strip() != NO_MOVE_VALUE:
            ws_target.cell(row=r, column=COL_B).value = v

    # 步骤 6：A 列不为空 → 剪切到「已出报告」
    print("步骤 6: A 不为空 → 「已出报告」…")
    move6, keep6 = [], []
    for r in range(2, ws_target.max_row + 1):
        row_vals = get_row(ws_target, r)
        if ws_target.cell(row=r, column=COL_A).value is not None:
            move6.append(row_vals)
        else:
            keep6.append(row_vals)
    for r in range(2, ws_target.max_row + 1):
        clear_row(ws_target, r)
    for i, row_vals in enumerate(keep6, start=2):
        set_row(ws_target, i, row_vals)
    append_rows(ws_reported, move6)
    print(f"  移动 {len(move6)} 行")

    # 步骤 7：「已出报告」B 含 "CIQ Classification" → 「共同分级」
    print("步骤 7: B 含 CIQ Classification → 「共同分级」…")
    move7, keep7 = [], []
    rep_last = last_data_row(ws_reported)
    for r in range(2, rep_last + 1):
        row_vals = get_row(ws_reported, r)
        b = ws_reported.cell(row=r, column=COL_B).value
        if b is not None and "CIQ Classification" in str(b):
            move7.append(row_vals)
        else:
            keep7.append(row_vals)
    for r in range(2, rep_last + 1):
        clear_row(ws_reported, r)
    for i, row_vals in enumerate(keep7, start=2):
        set_row(ws_reported, i, row_vals)
    append_rows(ws_joint, move7)
    print(f"  移动 {len(move7)} 行")

    # 步骤 8：「共同分级」H 不为空 → 「共同分级已完成」
    print("步骤 8: H 不为空 → 「共同分级已完成」…")
    move8, keep8 = [], []
    jnt_last = last_data_row(ws_joint)
    for r in range(2, jnt_last + 1):
        row_vals = get_row(ws_joint, r)
        h = ws_joint.cell(row=r, column=COL_H).value
        if h is not None:
            move8.append(row_vals)
        else:
            keep8.append(row_vals)
    for r in range(2, jnt_last + 1):
        clear_row(ws_joint, r)
    for i, row_vals in enumerate(keep8, start=2):
        set_row(ws_joint, i, row_vals)
    append_rows(ws_joint_done, move8)
    print(f"  移动 {len(move8)} 行")

    return wb2


# ── 主程序 ────────────────────────────────────────────────────────────────────

def main():
    print("脚本正在运行中，请稍候 …")
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        # 1. 登录
        print("=== 登录 SharePoint ===")
        login(page)

        # 2. 解析文件路径
        print("\n=== 解析文件路径 ===")
        file1_url = get_file_server_url(page, FILE1_GUID)
        file2_url = get_file_server_url(page, FILE2_GUID)
        print(f"  文件1: {file1_url}")
        print(f"  文件2: {file2_url}")

        # 3. 下载文件
        print("\n=== 下载文件 ===")
        f1_bytes = download_file_bytes(page, file1_url)
        f2_bytes = download_file_bytes(page, file2_url)
        print(f"  文件1 大小: {len(f1_bytes):,} bytes")
        print(f"  文件2 大小: {len(f2_bytes):,} bytes")

        # 4. 处理数据
        print("\n=== 处理 Excel 数据 ===")
        wb1 = openpyxl.load_workbook(io.BytesIO(f1_bytes))
        wb2 = openpyxl.load_workbook(io.BytesIO(f2_bytes))
        wb2 = process(wb1, wb2)

        # 5. 上传
        print("\n=== 上传文件 ===")
        buf = io.BytesIO()
        wb2.save(buf)
        buf.seek(0)
        upload_file_bytes(page, file2_url, buf.read())

        browser.close()

    print("\n全部完成！")


if __name__ == "__main__":
    main()
